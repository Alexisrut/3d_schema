"""Выгрузка задач и проблем в Excel.

Книга собирается в памяти и отдаётся потоком: файл получается небольшим
(тысячи строк — сотни килобайт), а временный файл на диске потребовал бы
уборки за собой.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .. import models, services
from ..deps import AccessibleProject, DbSession

router = APIRouter(prefix="/api/projects/{project_id}", tags=["export"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATUS_LABEL = {"todo": "В плане", "in_progress": "В работе", "done": "Готово"}

TASK_COLUMNS = [
    ("№", 6),
    ("Зона", 28),
    ("Задача", 34),
    ("Описание", 46),
    ("Статус", 14),
    ("Выполнено, %", 14),
    ("Бригады", 30),
    ("Создана", 18),
    ("В работе (ч:мм)", 16),
    ("Вложения", 30),
]

PROBLEM_COLUMNS = [
    ("№", 6),
    ("Зона", 28),
    ("Проблема", 34),
    ("Описание", 46),
    ("Состояние", 14),
    ("Бригады", 30),
    ("Создана", 18),
    ("Открыта (ч:мм)", 16),
    ("Вложения", 30),
]


@router.get("/export.xlsx")
def export_xlsx(project: AccessibleProject, db: DbSession):
    """Все задачи и проблемы проекта одной книгой: два листа."""
    snapshot = services.build_project_snapshot(db, project)
    now = datetime.now(timezone.utc)

    book = Workbook()
    tasks_sheet = book.active
    tasks_sheet.title = "Задачи"
    _write_header(tasks_sheet, TASK_COLUMNS)

    row_index = 2
    for sector in snapshot.sectors:
        brigades = ", ".join(b.name for b in sector.brigades)
        for task in sector.tasks:
            tasks_sheet.append([
                row_index - 1,
                sector.name,
                task.name,
                task.definition,
                STATUS_LABEL.get(_value(task.status), _value(task.status)),
                task.progress,
                brigades,
                _local(task.created_at),
                _elapsed(task.created_at, now),
                ", ".join(a.filename for a in task.attachments),
            ])
            row_index += 1

    problems_sheet = book.create_sheet("Проблемы")
    _write_header(problems_sheet, PROBLEM_COLUMNS)

    row_index = 2
    for sector in snapshot.sectors:
        brigades = ", ".join(b.name for b in sector.brigades)
        for problem in sector.problems:
            problems_sheet.append([
                row_index - 1,
                sector.name,
                problem.name,
                problem.definition,
                "Решена" if problem.is_resolved else "Открыта",
                brigades,
                _local(problem.created_at),
                _elapsed(problem.created_at, now),
                ", ".join(a.filename for a in problem.attachments),
            ])
            row_index += 1

    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)

    stamp = now.strftime("%Y-%m-%d")
    # Имя файла — только латиница и цифры: кириллица в filename ломает часть
    # клиентов, поэтому человекочитаемое имя идёт в filename* по RFC 5987.
    ascii_name = f"tasks-problems-{project.id}-{stamp}.xlsx"
    human = f"Задачи и проблемы — {project.name} — {stamp}.xlsx"
    disposition = (
        f"attachment; filename=\"{ascii_name}\"; "
        f"filename*=UTF-8''{_quote(human)}"
    )
    return StreamingResponse(
        buffer,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": disposition},
    )


def _write_header(sheet, columns: list[tuple[str, int]]) -> None:  # noqa: ANN001
    sheet.append([title for title, _ in columns])
    for index, (_, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(columns))}1"
    )


def _value(status: object) -> str:
    return str(getattr(status, "value", status))


def _local(moment: datetime) -> str:
    return _aware(moment).strftime("%d.%m.%Y %H:%M")


def _elapsed(created: datetime, now: datetime) -> str:
    """Сколько времени прошло с создания — часы и минуты, как в карточке."""
    delta = now - _aware(created)
    minutes = max(0, int(delta.total_seconds() // 60))
    return f"{minutes // 60}:{minutes % 60:02d}"


def _aware(moment: datetime) -> datetime:
    """SQLite отдаёт время без зоны — считаем его UTC, как и записывали."""
    return moment if moment.tzinfo else moment.replace(tzinfo=timezone.utc)


def _quote(value: str) -> str:
    from urllib.parse import quote

    return quote(value, safe="")
